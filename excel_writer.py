import openpyxl
import tkinter as tk
from tkinter import messagebox

def open_workbook(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        print(f"Workbook '{file_path}' loaded successfully.")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        print(f"Workbook '{file_path}' not found. Created a new one.")
    return workbook

def append_data_to_sheet(workbook, sheet_name, data):
    # Ensure all fields except "Check Number" have values
    for index, value in enumerate(data):
        if index != CHECK_NUMBER_INDEX and not value:  # Replace CHECK_NUMBER_INDEX with the actual index
            messagebox.showerror("Validation Error", f"Field at index {index} cannot be blank.")
            return False

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Convert data list to string for easier comparison
    data_str = "|".join([str(cell) for cell in data if cell])  # Handle blank values

    # Count duplicate entries
    duplicate_count = 0
    for row in sheet.iter_rows(values_only=True):
        row_str = "|".join([str(cell) for cell in row if cell])  # Handle blank values
        if row_str == data_str:
            duplicate_count += 1

    # Display a single message based on the count
    if duplicate_count > 0:
        match_word = "match" if duplicate_count == 1 else "matches"
        if not messagebox.askyesno("Duplicate Entry Found", f"{duplicate_count} exact {match_word} found in the Excel file. Do you want to add it anyway?"):
            return False

    # Append the data as a new row
    sheet.append(data)
    return True

def save_workbook(workbook, file_path):
    workbook.save(file_path)
    print(f"Workbook '{file_path}' saved successfully.")

def append_processed_data(file_path, sheet_name, data):
    workbook = open_workbook(file_path)
    if append_data_to_sheet(workbook, sheet_name, data):
        save_workbook(workbook, file_path)
        print("Data appended successfully.")
    else:
        print("Data append aborted due to duplicate entry.")

# Replace CHECK_NUMBER_INDEX with the actual index of the "Check Number" field in your data list
CHECK_NUMBER_INDEX = 4
